import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_mock_grades():
    """生成模拟成绩数据"""
    students = [
        ("张三", "202401001"), ("李四", "202401002"), ("王五", "202401003"),
        ("赵六", "202401004"), ("钱七", "202401005"), ("孙八", "202401006"),
        ("周九", "202401007"), ("吴十", "202401008"), ("郑十一", "202401009"),
        ("陈十二", "202401010"), ("刘十三", "202401011"), ("杨十四", "202401012"),
        ("朱十五", "202401013"), ("马十六", "202401014"), ("胡十七", "202401015"),
        ("郭十八", "202401016"), ("林十九", "202401017"), ("何二十", "202401018"),
        ("高二十一", "202401019"), ("梁二十二", "202401020"), ("黄二十三", "202401021"),
        ("周二十四", "202401022"), ("吴二十五", "202401023"), ("徐二十六", "202401024"),
        ("孙二十七", "202401025"), ("马二十八", "202401026"), ("朱二十九", "202401027"),
        ("胡三十", "202401028"), ("郭三十一", "202401029"), ("林三十二", "202401030"),
        ("何三十三", "202401031"), ("高三十四", "202401032"), ("梁三十五", "202401033"),
        ("黄三十六", "202401034"), ("周三十七", "202401035"), ("吴三十八", "202401036"),
        ("徐三十九", "202401037"), ("孙四十", "202401038"), ("马四十一", "202401039"),
        ("朱四十二", "202401040"), ("胡四十三", "202401041"), ("郭四十四", "202401042"),
        ("林四十五", "202401043")
    ]

    grades = []
    for name, no in students:
        # 生成符合正态分布的成绩
        usual = round(random.gauss(78, 12), 1)
        midterm = round(random.gauss(70, 15), 1)
        final = round(random.gauss(75, 14), 1)
        practice = round(random.gauss(82, 10), 1) if random.random() > 0.2 else None

        # 确保成绩在合理范围内
        usual = max(50, min(98, usual))
        midterm = max(35, min(98, midterm))
        final = max(40, min(99, final))
        if practice:
            practice = max(55, min(100, practice))

        grades.append({
            "student_name": name,
            "student_no": no,
            "usual_score": usual,
            "midterm_score": midterm,
            "final_score": final,
            "practice_score": practice
        })

    return grades

def create_grade_excel():
    """创建成绩Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "成绩单"

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="0891B2", end_color="0891B2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 标题行
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "2026年春季学期 市场营销 成绩表"
    title_cell.font = Font(bold=True, size=16, color="0891B2")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # 表头
    headers = ["学生姓名", "学号", "平时分", "期中分", "期末分", "实验分"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    ws.row_dimensions[2].height = 25

    # 生成数据
    grades = generate_mock_grades()
    for row_idx, grade in enumerate(grades, 1):
        row = row_idx + 2

        ws.cell(row=row, column=1, value=grade["student_name"]).border = thin_border
        ws.cell(row=row, column=2, value=grade["student_no"]).border = thin_border
        ws.cell(row=row, column=3, value=grade["usual_score"]).border = thin_border
        ws.cell(row=row, column=4, value=grade["midterm_score"]).border = thin_border
        ws.cell(row=row, column=5, value=grade["final_score"]).border = thin_border

        practice_cell = ws.cell(row=row, column=6, value=grade["practice_score"])
        practice_cell.border = thin_border
        if grade["practice_score"]:
            practice_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")

        # 根据期末成绩着色
        final_score = grade["final_score"]
        score_cell = ws.cell(row=row, column=5)
        if final_score >= 90:
            score_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        elif final_score < 60:
            score_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    # 设置列宽
    column_widths = [12, 12, 10, 10, 10, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # 统计行
    stats_row = len(grades) + 3
    ws.cell(row=stats_row, column=1, value="📊 统计信息").font = Font(bold=True)
    ws.cell(row=stats_row + 1, column=1, value="班级人数：45人")
    ws.cell(row=stats_row + 2, column=1, value="考试类型：闭卷")
    ws.cell(row=stats_row + 3, column=1, value="考试时间：2026年1月15日")

    # 保存
    output_path = "/Users/salmon/Desktop/Xunfei project/AI小商/模拟成绩表_市场营销_2026春季.xlsx"
    wb.save(output_path)
    print(f"✅ 成绩表已生成: {output_path}")
    return output_path

if __name__ == "__main__":
    create_grade_excel()
