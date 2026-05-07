"""
成绩管理服务

提供成绩上传解析、AI分析报告生成、成绩统计、导出等功能
"""
import logging
import json
import re
from io import BytesIO
from typing import List, Dict, Optional, Any
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.utils.errors import ValidationException, ThirdPartyException

logger = logging.getLogger("grade")

# AI分析报告Prompt模板
REPORT_PROMPT = """你是一个专业的教学分析师。请根据以下成绩数据生成一份详细的班级成绩分析报告。

课程：{course_name}
学期：{semester}
班级人数：{student_count}

成绩统计数据：
- 平均分：{avg_score}
- 及格率：{pass_rate}
- 最高分：{max_score}
- 最低分：{min_score}

成绩分布：
{distribution}

各项成绩平均分：
- 平时分平均：{usual_avg}
- 期中分平均：{midterm_avg}
- 期末分平均：{final_avg}
- 实验分平均：{practice_avg}

请生成以下JSON格式的分析报告：
{{
  "summary": "班级整体表现概述（100字以内）",
  "high_performers": ["高分学生名单，最多5人"],
  "needs_attention": ["需要关注的学生名单，最多5人"],
  "exam_analysis": {{
    "difficulty": 难度系数(0-1之间，可用max_score和avg_score推算),
    "difficulty_text": "难度描述（很难/较难/适中/较易/很容易）",
    "discrimination": 区分度(0-1之间，可用分数标准差推算),
    "discrimination_text": "区分度描述（很差/较差/良好/优秀）"
  }},
  "suggestions": "教学建议（3条以内，50字以内）"
}}

注意：只返回JSON，不要有其他内容。"""


class GradeService:
    """成绩管理服务"""

    # Excel列名映射（支持多种命名）
    COLUMN_MAPPING = {
        'student_name': ['学生姓名', '姓名', 'name', 'student_name', '学生名字'],
        'student_no': ['学号', 'student_no', 'student number', '编号', 'student_number'],
        'usual_score': ['平时分', '平时成绩', 'usual', 'usual_score', '平时'],
        'midterm_score': ['期中分', '期中成绩', 'midterm', 'midterm_score', '期中'],
        'final_score': ['期末分', '期末成绩', 'final', 'final_score', '期末'],
        'practice_score': ['实验分', '实验成绩', '实践分', '实践成绩', 'practice', 'practice_score', '实验', '实践'],
    }

    # 默认权重
    DEFAULT_WEIGHTS = {"usual": 0.4, "midterm": 0.2, "final": 0.4, "practice": 0.0}

    def __init__(self):
        self.logger = logger

    def _map_column(self, header: str) -> Optional[str]:
        """将表头列名映射到标准字段名"""
        if not header:
            return None
        header_lower = str(header).lower().strip()
        for standard_name, variants in self.COLUMN_MAPPING.items():
            if header_lower in [v.lower() for v in variants]:
                self.logger.debug(f"[Grade] Column mapped: '{header}' -> '{standard_name}'")
                return standard_name
        return None

    def _find_header_row(self, ws) -> int:
        """查找包含成绩列名的行（跳过标题行）"""
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            # 检查这行是否包含"学生姓名"或"姓名"
            row_values = [str(v).lower() if v else "" for v in row]
            has_student_name = any('学生姓名' in v or '姓名' in v or v == 'name' for v in row_values)
            has_final = any('期末' in v or 'final' in v for v in row_values)

            if has_student_name and has_final:
                self.logger.info(f"[Grade] Header row found at row {row_idx}: {list(row)}")
                return row_idx

        # 没找到，返回第1行
        self.logger.warning(f"[Grade] Could not find header row, using row 1")
        return 1

    async def parse_excel(self, file_data: bytes, weights: Optional[Dict[str, float]] = None) -> List[Dict[str, Any]]:
        """
        解析Excel文件

        Args:
            file_data: Excel文件二进制数据
            weights: 权重配置，默认使用DEFAULT_WEIGHTS

        Returns:
            成绩列表 [{student_name, usual_score, midterm_score, final_score, practice_score, total_score}, ...]
        """
        self.logger.info(f"[Grade] Parsing excel, size: {len(file_data)} bytes")

        if weights is None:
            weights = self.DEFAULT_WEIGHTS

        try:
            wb = load_workbook(BytesIO(file_data))
            ws = wb.active

            # 查找表头行
            header_row = self._find_header_row(ws)
            self.logger.info(f"[Grade] Using header row: {header_row}")

            # 读取表头行（values_only=True返回的是值元组，不是Cell对象）
            headers = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
            self.logger.info(f"[Grade] Raw headers: {headers}")

            # 建立列名映射
            column_map = {}
            for i, header in enumerate(headers):
                mapped = self._map_column(header)
                if mapped:
                    column_map[mapped] = i

            self.logger.info(f"[Grade] Column map: {column_map}")

            # 检查必需列
            required_cols = ['student_name', 'final_score']
            missing_cols = [col for col in required_cols if col not in column_map]
            if missing_cols:
                self.logger.error(f"[Grade] Missing required columns: {missing_cols}")
                raise ValidationException(
                    message="Excel格式不正确",
                    details={
                        "missing_column": missing_cols[0],
                        "hint": "请确保包含：学生姓名、期末分（必填），平时分、期中分、实验分（选填）",
                        "found_headers": headers
                    }
                )

            # 解析数据行
            items = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_row=None, values_only=True), start=header_row + 1):
                item = {}

                # 映射各字段
                for field_name, col_idx in column_map.items():
                    value = row[col_idx] if col_idx < len(row) else None
                    item[field_name] = value

                # 跳过空行
                if not item.get('student_name'):
                    self.logger.debug(f"[Grade] Skipping empty row {row_idx}")
                    continue

                # 计算总评
                total = 0.0
                weight_sum = 0.0

                for key, weight in weights.items():
                    score = item.get(f"{key}_score")
                    if score is not None and weight > 0:
                        try:
                            total += float(score) * weight
                            weight_sum += weight
                        except (ValueError, TypeError) as e:
                            self.logger.warning(f"[Grade] Could not parse score '{score}' for {item.get('student_name')}: {e}")

                item["total_score"] = round(total / weight_sum if weight_sum > 0 else 0, 2)
                item["status"] = "normal"
                items.append(item)

                self.logger.debug(f"[Grade] Row {row_idx}: {item.get('student_name')}, total={item['total_score']}")

            self.logger.info(f"[Grade] Parsed {len(items)} items successfully")

            if not items:
                raise ValidationException(
                    message="成绩表为空",
                    details={"hint": "请确保Excel包含学生成绩数据"}
                )

            return items

        except ValidationException:
            raise
        except Exception as e:
            self.logger.error(f"[Grade] Parse failed: {str(e)}", exc_info=True)
            raise ValidationException(
                message="Excel解析失败",
                details={"error": str(e), "error_type": type(e).__name__}
            )

    async def calculate_grades(self, items: List[Dict[str, Any]], weights: Optional[Dict[str, float]] = None) -> List[Dict[str, Any]]:
        """
        计算成绩（含排名）

        Args:
            items: 原始成绩列表
            weights: 权重配置

        Returns:
            含总评和排名的成绩列表
        """
        if weights is None:
            weights = self.DEFAULT_WEIGHTS

        self.logger.info(f"[Grade] Calculating {len(items)} grades with weights: {weights}")

        # 计算总评
        for item in items:
            total = 0.0
            weight_sum = 0.0

            for key, weight in weights.items():
                score = item.get(f"{key}_score")
                if score is not None and weight > 0:
                    try:
                        total += float(score) * weight
                        weight_sum += weight
                    except (ValueError, TypeError):
                        pass

            item["total_score"] = round(total / weight_sum if weight_sum > 0 else 0, 2)

        # 计算排名（按总评降序）
        sorted_items = sorted(items, key=lambda x: x.get("total_score", 0), reverse=True)
        for rank, item in enumerate(sorted_items, 1):
            item["rank"] = rank

        self.logger.info(f"[Grade] Ranking calculated for {len(items)} students")
        return items

    async def calculate_stats(self, items: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        计算统计数据

        Args:
            items: 成绩列表（含total_score）

        Returns:
            {
                basic: {total_students, avg_score, pass_rate, max_score, min_score},
                distribution: [{range, count, percentage}, ...],
                by_component: {usual: {avg, max, min}, ...}
            }
        """
        if not items:
            return {
                "basic": {"total_students": 0, "avg_score": 0, "pass_rate": 0, "max_score": 0, "min_score": 0},
                "distribution": [],
                "by_component": None
            }

        total_scores = [item.get("total_score", 0) for item in items if item.get("total_score") is not None]
        total_students = len(items)

        # 基础统计
        avg_score = round(sum(total_scores) / len(total_scores), 2) if total_scores else 0
        pass_rate = round(sum(1 for s in total_scores if s >= 60) / len(total_scores), 4) if total_scores else 0
        max_score = max(total_scores) if total_scores else 0
        min_score = min(total_scores) if total_scores else 0

        basic = {
            "total_students": total_students,
            "avg_score": avg_score,
            "pass_rate": pass_rate,
            "max_score": max_score,
            "min_score": min_score
        }

        # 成绩分布
        ranges = ["<60", "60-70", "70-80", "80-90", ">90"]
        distribution = []
        for r in ranges:
            if r == "<60":
                count = sum(1 for s in total_scores if s < 60)
            elif r == ">90":
                count = sum(1 for s in total_scores if s >= 90)
            else:
                low, high = map(int, r.split("-"))
                count = sum(1 for s in total_scores if low <= s < high)
            percentage = round(count / total_students * 100, 1) if total_students > 0 else 0
            distribution.append({"range": r, "count": count, "percentage": percentage})

        # 各单项统计
        components = ["usual_score", "midterm_score", "final_score", "practice_score"]
        component_names = ["usual", "midterm", "final", "practice"]
        by_component = {}

        for comp_key, comp_name in zip(components, component_names):
            scores = [item.get(comp_key) for item in items if item.get(comp_key) is not None]
            if scores:
                valid_scores = [s for s in scores if isinstance(s, (int, float))]
                if valid_scores:
                    by_component[comp_name] = {
                        "avg": round(sum(valid_scores) / len(valid_scores), 2),
                        "max": max(valid_scores),
                        "min": min(valid_scores)
                    }

        self.logger.info(f"[Grade] Stats calculated: avg={avg_score}, pass_rate={pass_rate}, students={total_students}")

        return {
            "basic": basic,
            "distribution": distribution,
            "by_component": by_component if by_component else None
        }

    async def generate_ai_report(
        self,
        course_name: str,
        semester: str,
        items: List[Dict[str, Any]],
        stats: Dict[str, Any]
    ) -> Dict[str, Any]:
        """
        生成AI分析报告（调用星火大模型）

        Args:
            course_name: 课程名称
            semester: 学期
            items: 成绩列表
            stats: 统计数据

        Returns:
            {
                summary: str,
                high_performers: [str],
                needs_attention: [str],
                exam_analysis: {difficulty, difficulty_text, discrimination, discrimination_text},
                suggestions: str
            }
        """
        self.logger.info(f"[Grade] Generating AI report for {course_name}")

        basic = stats.get("basic", {})
        distribution = stats.get("distribution", [])
        by_component = stats.get("by_component", {})

        # 整理分布数据
        dist_str = "\n".join([f"{d['range']}: {d['count']}人 ({d['percentage']}%)" for d in distribution])

        # 获取各单项均分
        usual_avg = by_component.get("usual", {}).get("avg", "无数据")
        midterm_avg = by_component.get("midterm", {}).get("avg", "无数据")
        final_avg = by_component.get("final", {}).get("avg", "无数据")
        practice_avg = by_component.get("practice", {}).get("avg", "无数据")

        # 构建Prompt
        prompt = REPORT_PROMPT.format(
            course_name=course_name,
            semester=semester or "未知",
            student_count=basic.get("total_students", 0),
            avg_score=basic.get("avg_score", 0),
            pass_rate=f"{basic.get('pass_rate', 0) * 100:.1f}%",
            max_score=basic.get("max_score", 0),
            min_score=basic.get("min_score", 0),
            distribution=dist_str,
            usual_avg=usual_avg,
            midterm_avg=midterm_avg,
            final_avg=final_avg,
            practice_avg=practice_avg
        )

        self.logger.info(f"[Grade] Report prompt length: {len(prompt)}")

        try:
            # 调用星火大模型
            from app.services.xinghuo_service import xinghuo_service

            messages = [{"role": "user", "content": prompt}]
            response = await xinghuo_service.chat_completion(
                messages=messages,
                user_id="grade_analyzer",
                temperature=0.5,
                max_tokens=2048
            )

            self.logger.info(f"[Grade] AI response length: {len(response)}")
            self.logger.debug(f"[Grade] AI response preview: {response[:300]}")

            # 解析JSON响应
            report_data = self._parse_json_response(response)

            self.logger.info(f"[Grade] AI report parsed successfully")

            return {
                "course_name": course_name,
                "semester": semester,
                "summary": report_data.get("summary", ""),
                "high_performers": report_data.get("high_performers", []),
                "needs_attention": report_data.get("needs_attention", []),
                "exam_analysis": {
                    "difficulty": report_data.get("exam_analysis", {}).get("difficulty", 0.5),
                    "difficulty_text": report_data.get("exam_analysis", {}).get("difficulty_text", "适中"),
                    "discrimination": report_data.get("exam_analysis", {}).get("discrimination", 0.3),
                    "discrimination_text": report_data.get("exam_analysis", {}).get("discrimination_text", "良好")
                },
                "suggestions": report_data.get("suggestions", "")
            }

        except ValidationException:
            raise
        except ThirdPartyException:
            # 星火API错误，记录并返回默认报告
            self.logger.warning(f"[Grade] AI report generation failed, using default report")
            return await self._generate_default_report(course_name, semester, items, stats)
        except Exception as e:
            self.logger.error(f"[Grade] AI report generation failed: {str(e)}", exc_info=True)
            return await self._generate_default_report(course_name, semester, items, stats)

    def _parse_json_response(self, response: str) -> Dict[str, Any]:
        """解析星火返回的JSON响应"""
        # 尝试提取JSON代码块
        json_str = response.strip()

        # 处理 markdown 代码块
        if "```json" in json_str:
            json_str = json_str.split("```json")[1].split("```")[0]
        elif "```" in json_str:
            json_str = json_str.split("```")[1].split("```")[0]

        json_str = json_str.strip()

        # 尝试直接解析
        try:
            return json.loads(json_str)
        except json.JSONDecodeError:
            pass

        # 尝试用正则提取JSON对象
        json_match = re.search(r'\{[\s\S]*\}', json_str)
        if json_match:
            try:
                return json.loads(json_match.group())
            except json.JSONDecodeError:
                pass

        # 尝试清理常见的格式问题
        # 移除尾部的多余字符
        for i in range(len(json_str), 0, -1):
            try:
                return json.loads(json_str[:i])
            except json.JSONDecodeError:
                continue

        raise ValidationException(
            message="AI返回格式解析失败",
            details={"response_preview": response[:500]}
        )

    async def _generate_default_report(
        self,
        course_name: str,
        semester: str,
        items: List[Dict[str, Any]],
        stats: Dict[str, Any]
    ) -> Dict[str, Any]:
        """生成默认报告（当AI调用失败时）"""
        self.logger.info(f"[Grade] Generating default report")
        basic = stats.get("basic", {})

        # 简单计算难度和区分度
        avg = basic.get("avg_score", 0)
        max_s = basic.get("max_score", 100)
        min_s = basic.get("min_score", 0)

        # 难度：基于平均分估算 (平均分高则难度低)
        difficulty = 1 - (avg / 100) if avg else 0.5

        # 区分度：使用分数标准差估算
        total_scores = [item.get("total_score", 0) for item in items if item.get("total_score") is not None]
        if len(total_scores) > 1:
            import statistics
            std_dev = statistics.stdev(total_scores)
            discrimination = min(std_dev / 50, 1.0)  # 标准差/50作为区分度
        else:
            discrimination = 0.3

        if difficulty < 0.3:
            difficulty_text = "很容易"
        elif difficulty < 0.5:
            difficulty_text = "较易"
        elif difficulty < 0.7:
            difficulty_text = "适中"
        elif difficulty < 0.85:
            difficulty_text = "较难"
        else:
            difficulty_text = "很难"

        discrimination_text = "优秀" if discrimination > 0.4 else "良好" if discrimination > 0.25 else "较差"

        # 找出高分和低分学生
        sorted_items = sorted(items, key=lambda x: x.get("total_score", 0), reverse=True)
        high_performers = [item.get("student_name", "") for item in sorted_items[:5] if item.get("student_name")]
        needs_attention = [item.get("student_name", "") for item in sorted_items[-5:] if item.get("student_name")]

        summary = f"本学期《{course_name}》班级整体表现{'良好' if avg >= 70 else '一般'}，平均分{avg}分，及格率{basic.get('pass_rate', 0) * 100:.1f}%。"
        if avg >= 70:
            summary += "大部分学生掌握了课程内容。"

        return {
            "course_name": course_name,
            "semester": semester,
            "summary": summary,
            "high_performers": high_performers,
            "needs_attention": needs_attention,
            "exam_analysis": {
                "difficulty": round(difficulty, 2),
                "difficulty_text": difficulty_text,
                "discrimination": round(discrimination, 2),
                "discrimination_text": discrimination_text
            },
            "suggestions": f"1. {'继续保持当前教学方式' if avg >= 70 else '加强基础内容讲解'}\n2. 关注需要帮助的学生\n3. 适当调整教学方法"
        }

    async def export_excel(
        self,
        course_name: str,
        class_name: str,
        semester: str,
        items: List[Dict[str, Any]],
        weights: Dict[str, float]
    ) -> bytes:
        """
        导出成绩Excel

        Args:
            course_name: 课程名称
            class_name: 班级名称
            semester: 学期
            items: 成绩列表
            weights: 权重配置

        Returns:
            Excel文件二进制数据
        """
        self.logger.info(f"[Grade] Exporting excel for {course_name}, {len(items)} items")

        wb = Workbook()
        ws = wb.active
        ws.title = "成绩单"

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0891B2", end_color="0891B2", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 标题行
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"{course_name} - {class_name or '未知班级'} - {semester or '未知学期'} 成绩单"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(start_color="22D3EE", end_color="22D3EE", fill_type="solid")

        ws.row_dimensions[1].height = 30

        # 表头
        headers = ["排名", "学生姓名", "学号", "平时分", "期中分", "期末分", "实验分", "总评"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        ws.row_dimensions[2].height = 25

        # 数据行
        for row_idx, item in enumerate(items, 1):
            row = row_idx + 2

            ws.cell(row=row, column=1, value=item.get("rank", "-")).border = thin_border
            ws.cell(row=row, column=2, value=item.get("student_name", "")).border = thin_border
            ws.cell(row=row, column=3, value=item.get("student_no", "")).border = thin_border
            ws.cell(row=row, column=4, value=item.get("usual_score", "-")).border = thin_border
            ws.cell(row=row, column=5, value=item.get("midterm_score", "-")).border = thin_border
            ws.cell(row=row, column=6, value=item.get("final_score", "-")).border = thin_border
            ws.cell(row=row, column=7, value=item.get("practice_score", "-")).border = thin_border
            ws.cell(row=row, column=8, value=item.get("total_score", "-")).border = thin_border

            # 及格/不及格颜色标记
            total = item.get("total_score", 0)
            if isinstance(total, (int, float)):
                if total < 60:
                    ws.cell(row=row, column=8).fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                elif total >= 90:
                    ws.cell(row=row, column=8).fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")

        # 统计行
        stats_row = len(items) + 3
        total_scores = [item.get("total_score", 0) for item in items if isinstance(item.get("total_score"), (int, float))]
        if total_scores:
            ws.cell(row=stats_row, column=1, value="统计").font = Font(bold=True)
            ws.cell(row=stats_row, column=6, value=f"平均分: {sum(total_scores)/len(total_scores):.1f}")
            ws.cell(row=stats_row, column=7, value=f"最高分: {max(total_scores)}")
            ws.cell(row=stats_row, column=8, value=f"最低分: {min(total_scores)}")

        # 设置列宽
        column_widths = [8, 12, 12, 10, 10, 10, 10, 10]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # 保存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        self.logger.info(f"[Grade] Excel exported successfully, size: {output.getbuffer().nbytes} bytes")
        return output.getvalue()


# 全局单例
grade_service = GradeService()
